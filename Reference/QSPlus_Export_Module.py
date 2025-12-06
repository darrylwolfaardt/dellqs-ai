"""
QSPlus Export Module for DELLQS-AI
====================================

This module provides functionality to export Bill of Quantities (BOQ) data
in formats compatible with QSPlus estimating software.

Supported Export Formats:
- Excel (.xlsx) - Standard BOQ format
- CSV (.csv) - Simple comma-separated format
- XML (.xml) - Structured data format

Author: DELLQS-AI
Date: 2025-12-06
Version: 1.0
"""

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Union
import xml.etree.ElementTree as ET
from xml.dom import minidom


class BOQItem:
    """Represents a single item in the Bill of Quantities"""

    def __init__(
        self,
        item_number: str,
        description: str,
        unit: str,
        quantity: float = 0.0,
        rate: float = 0.0,
        amount: float = 0.0,
        section: str = "",
        subsection: str = "",
        calculation_notes: str = "",
        reference_drawing: str = "",
        measurement_rule: str = ""
    ):
        self.item_number = item_number
        self.description = description
        self.unit = unit
        self.quantity = quantity
        self.rate = rate
        self.amount = amount if amount > 0 else (quantity * rate)
        self.section = section
        self.subsection = subsection
        self.calculation_notes = calculation_notes
        self.reference_drawing = reference_drawing
        self.measurement_rule = measurement_rule

    def to_dict(self) -> Dict:
        """Convert BOQ item to dictionary"""
        return {
            'Item Number': self.item_number,
            'Section': self.section,
            'Subsection': self.subsection,
            'Description': self.description,
            'Unit': self.unit,
            'Quantity': self.quantity,
            'Rate': self.rate,
            'Amount': self.amount,
            'Calculation Notes': self.calculation_notes,
            'Reference Drawing': self.reference_drawing,
            'Measurement Rule': self.measurement_rule
        }


class QSPlusExporter:
    """Main class for exporting BOQ data to QSPlus-compatible formats"""

    def __init__(self, project_name: str = "", project_number: str = ""):
        self.project_name = project_name
        self.project_number = project_number
        self.items: List[BOQItem] = []
        self.metadata = {
            'created_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'software': 'DELLQS-AI',
            'version': '1.0'
        }

    def add_item(self, item: BOQItem):
        """Add a BOQ item to the export"""
        self.items.append(item)

    def add_items(self, items: List[BOQItem]):
        """Add multiple BOQ items to the export"""
        self.items.extend(items)

    def clear_items(self):
        """Clear all items from the export"""
        self.items = []

    def export_to_csv(self, output_path: str, include_calculations: bool = True) -> str:
        """
        Export BOQ to CSV format

        Args:
            output_path: Path where CSV file will be saved
            include_calculations: Whether to include calculation notes and references

        Returns:
            Path to the created CSV file
        """
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        # Define headers based on include_calculations flag
        if include_calculations:
            headers = [
                'Item Number', 'Section', 'Subsection', 'Description',
                'Unit', 'Quantity', 'Rate', 'Amount',
                'Calculation Notes', 'Reference Drawing', 'Measurement Rule'
            ]
        else:
            headers = [
                'Item Number', 'Section', 'Subsection', 'Description',
                'Unit', 'Quantity', 'Rate', 'Amount'
            ]

        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers)

            # Write header section
            writer.writerow({headers[0]: f'Project: {self.project_name}'})
            writer.writerow({headers[0]: f'Project Number: {self.project_number}'})
            writer.writerow({headers[0]: f'Date: {self.metadata["created_date"]}'})
            writer.writerow({headers[0]: f'Software: {self.metadata["software"]}'})
            writer.writerow({})  # Blank line

            # Write column headers
            writer.writeheader()

            # Write data rows
            for item in self.items:
                row_data = item.to_dict()
                if not include_calculations:
                    # Remove calculation-related fields
                    row_data = {k: v for k, v in row_data.items() if k in headers}
                writer.writerow(row_data)

            # Write summary section
            total_amount = sum(item.amount for item in self.items)
            writer.writerow({})  # Blank line
            writer.writerow({
                'Description': 'TOTAL',
                'Amount': f'{total_amount:.2f}'
            })

        return str(output_file)

    def export_to_excel(self, output_path: str, include_calculations: bool = True) -> str:
        """
        Export BOQ to Excel format (.xlsx)

        Args:
            output_path: Path where Excel file will be saved
            include_calculations: Whether to include calculation notes and references

        Returns:
            Path to the created Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        # Create workbook and select active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bill of Quantities"

        # Define styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        title_font = Font(bold=True, size=14)
        section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        section_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write project information
        row = 1
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = f'BILL OF QUANTITIES - {self.project_name}'
        cell.font = title_font

        row += 1
        ws[f'A{row}'] = f'Project Number:'
        ws[f'B{row}'] = self.project_number

        row += 1
        ws[f'A{row}'] = f'Date:'
        ws[f'B{row}'] = self.metadata['created_date']

        row += 1
        ws[f'A{row}'] = f'Generated by:'
        ws[f'B{row}'] = self.metadata['software']

        row += 2  # Blank line

        # Write column headers
        if include_calculations:
            headers = [
                'Item No.', 'Section', 'Subsection', 'Description',
                'Unit', 'Quantity', 'Rate', 'Amount',
                'Calculation Notes', 'Reference', 'Rule'
            ]
            col_widths = [10, 20, 20, 50, 10, 12, 12, 15, 40, 15, 15]
        else:
            headers = [
                'Item No.', 'Section', 'Subsection', 'Description',
                'Unit', 'Quantity', 'Rate', 'Amount'
            ]
            col_widths = [10, 20, 20, 50, 10, 12, 12, 15]

        for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        row += 1
        data_start_row = row

        # Write data rows
        current_section = ""
        for item in self.items:
            # Add section header if section changes
            if item.section and item.section != current_section:
                ws.merge_cells(f'A{row}:{openpyxl.utils.get_column_letter(len(headers))}{row}')
                cell = ws[f'A{row}']
                cell.value = item.section
                cell.font = section_font
                cell.fill = section_fill
                cell.alignment = Alignment(horizontal='left', vertical='center')
                row += 1
                current_section = item.section

            # Write item data
            data = [
                item.item_number,
                item.section,
                item.subsection,
                item.description,
                item.unit,
                item.quantity,
                item.rate,
                item.amount
            ]

            if include_calculations:
                data.extend([
                    item.calculation_notes,
                    item.reference_drawing,
                    item.measurement_rule
                ])

            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border

                # Format numbers
                if col in [6, 7, 8]:  # Quantity, Rate, Amount columns
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')

            row += 1

        # Write summary section
        row += 1
        total_amount = sum(item.amount for item in self.items)

        ws.merge_cells(f'A{row}:{openpyxl.utils.get_column_letter(len(headers)-1)}{row}')
        cell = ws[f'A{row}']
        cell.value = 'TOTAL'
        cell.font = section_font
        cell.alignment = Alignment(horizontal='right')

        cell = ws.cell(row=row, column=len(headers))
        cell.value = total_amount
        cell.font = section_font
        cell.number_format = '#,##0.00'
        cell.border = border

        # Freeze panes (freeze header row)
        ws.freeze_panes = f'A{data_start_row}'

        # Save workbook
        wb.save(output_file)
        return str(output_file)

    def export_to_xml(self, output_path: str, include_calculations: bool = True) -> str:
        """
        Export BOQ to XML format

        Args:
            output_path: Path where XML file will be saved
            include_calculations: Whether to include calculation notes and references

        Returns:
            Path to the created XML file
        """
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        # Create root element
        root = ET.Element('BillOfQuantities')

        # Add project information
        project_info = ET.SubElement(root, 'ProjectInformation')
        ET.SubElement(project_info, 'ProjectName').text = self.project_name
        ET.SubElement(project_info, 'ProjectNumber').text = self.project_number
        ET.SubElement(project_info, 'CreatedDate').text = self.metadata['created_date']
        ET.SubElement(project_info, 'Software').text = self.metadata['software']
        ET.SubElement(project_info, 'Version').text = self.metadata['version']

        # Add items
        items_element = ET.SubElement(root, 'Items')

        for item in self.items:
            item_element = ET.SubElement(items_element, 'Item')
            ET.SubElement(item_element, 'ItemNumber').text = item.item_number
            ET.SubElement(item_element, 'Section').text = item.section
            ET.SubElement(item_element, 'Subsection').text = item.subsection
            ET.SubElement(item_element, 'Description').text = item.description
            ET.SubElement(item_element, 'Unit').text = item.unit
            ET.SubElement(item_element, 'Quantity').text = str(item.quantity)
            ET.SubElement(item_element, 'Rate').text = str(item.rate)
            ET.SubElement(item_element, 'Amount').text = str(item.amount)

            if include_calculations:
                ET.SubElement(item_element, 'CalculationNotes').text = item.calculation_notes
                ET.SubElement(item_element, 'ReferenceDrawing').text = item.reference_drawing
                ET.SubElement(item_element, 'MeasurementRule').text = item.measurement_rule

        # Add summary
        summary = ET.SubElement(root, 'Summary')
        total_amount = sum(item.amount for item in self.items)
        ET.SubElement(summary, 'TotalItems').text = str(len(self.items))
        ET.SubElement(summary, 'TotalAmount').text = f'{total_amount:.2f}'

        # Pretty print XML
        xml_str = minidom.parseString(ET.tostring(root)).toprettyxml(indent="  ")

        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(xml_str)

        return str(output_file)

    def export_to_json(self, output_path: str, include_calculations: bool = True) -> str:
        """
        Export BOQ to JSON format

        Args:
            output_path: Path where JSON file will be saved
            include_calculations: Whether to include calculation notes and references

        Returns:
            Path to the created JSON file
        """
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        # Build JSON structure
        data = {
            'project_information': {
                'project_name': self.project_name,
                'project_number': self.project_number,
                'created_date': self.metadata['created_date'],
                'software': self.metadata['software'],
                'version': self.metadata['version']
            },
            'items': []
        }

        for item in self.items:
            item_data = item.to_dict()
            if not include_calculations:
                # Remove calculation-related fields
                item_data.pop('Calculation Notes', None)
                item_data.pop('Reference Drawing', None)
                item_data.pop('Measurement Rule', None)
            data['items'].append(item_data)

        # Add summary
        total_amount = sum(item.amount for item in self.items)
        data['summary'] = {
            'total_items': len(self.items),
            'total_amount': round(total_amount, 2)
        }

        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        return str(output_file)


# Convenience functions for quick exports
def export_boq_to_qsplus(
    items: List[BOQItem],
    output_path: str,
    project_name: str = "",
    project_number: str = "",
    format: str = "excel",
    include_calculations: bool = True
) -> str:
    """
    Quick export function for BOQ data to QSPlus-compatible format

    Args:
        items: List of BOQItem objects
        output_path: Path where file will be saved
        project_name: Name of the project
        project_number: Project reference number
        format: Export format ('excel', 'csv', 'xml', 'json')
        include_calculations: Whether to include calculation notes

    Returns:
        Path to the created file
    """
    exporter = QSPlusExporter(project_name, project_number)
    exporter.add_items(items)

    format = format.lower()
    if format == 'excel' or format == 'xlsx':
        return exporter.export_to_excel(output_path, include_calculations)
    elif format == 'csv':
        return exporter.export_to_csv(output_path, include_calculations)
    elif format == 'xml':
        return exporter.export_to_xml(output_path, include_calculations)
    elif format == 'json':
        return exporter.export_to_json(output_path, include_calculations)
    else:
        raise ValueError(f"Unsupported format: {format}. Use 'excel', 'csv', 'xml', or 'json'")


# Example usage
if __name__ == "__main__":
    # Example: Create some sample BOQ items
    sample_items = [
        BOQItem(
            item_number="1.1",
            section="FOUNDATIONS",
            subsection="Site Preparation",
            description="Clear site - Measure 2m beyond perimeter of building",
            unit="m2",
            quantity=150.5,
            rate=25.00,
            measurement_rule="DQSRule",
            calculation_notes="Building perimeter 10m x 12m + 2m margin"
        ),
        BOQItem(
            item_number="1.2",
            section="FOUNDATIONS",
            subsection="Site Preparation",
            description="Strip topsoil - Measure 1m beyond perimeter of building",
            unit="m2",
            quantity=132.0,
            rate=30.00,
            measurement_rule="DQSRule",
            calculation_notes="Building perimeter 10m x 12m + 1m margin"
        ),
        BOQItem(
            item_number="1.3",
            section="FOUNDATIONS",
            subsection="Excavation",
            description="Excavate for surface trenches - 700mm wide x 1500mm deep",
            unit="m3",
            quantity=15.4,
            rate=85.00,
            measurement_rule="DQSRule",
            calculation_notes="Perimeter 44m x 0.7m width x 1.5m depth"
        ),
    ]

    # Export to different formats
    exporter = QSPlusExporter(
        project_name="Sample Building Project",
        project_number="PROJ-2025-001"
    )
    exporter.add_items(sample_items)

    # Export to Excel
    excel_file = exporter.export_to_excel("sample_boq.xlsx")
    print(f"Excel export created: {excel_file}")

    # Export to CSV
    csv_file = exporter.export_to_csv("sample_boq.csv")
    print(f"CSV export created: {csv_file}")

    # Export to XML
    xml_file = exporter.export_to_xml("sample_boq.xml")
    print(f"XML export created: {xml_file}")

    # Export to JSON
    json_file = exporter.export_to_json("sample_boq.json")
    print(f"JSON export created: {json_file}")
