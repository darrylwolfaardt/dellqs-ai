"""
Cost Agent Implementation.

The Cost Agent applies rates to quantities, adjusts for regional factors,
and produces priced documents. It:
1. Receives measured quantities from the Measure Agent
2. Applies rates from the rate database
3. Adjusts for regional factors and market conditions
4. Calculates preliminaries, contingencies, and fees
5. Produces priced BOQ and cost summary

Outputs:
- priced_boq.xlsx: Bill of quantities with rates applied
- cost_summary.md: High-level cost breakdown
- pricing_assumptions.md: Rate sources and adjustment logic
- cost_risk_register.md: Cost risks and mitigation
"""

import asyncio
import json
import logging
import time
import uuid
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from enum import Enum
from pathlib import Path
from typing import Any, Optional

from .measure import (
    MeasureResult,
    QuantityItem,
    ElementGroup,
    ElementCategory,
    UnitOfMeasure,
    MeasurementStandard,
)

logger = logging.getLogger(__name__)


class ProjectStage(Enum):
    """Project stage affects pricing detail level."""
    FEASIBILITY = "feasibility"  # Order of magnitude estimate
    CONCEPT = "concept"  # Budget estimate
    DEVELOPED = "developed"  # Cost plan
    TECHNICAL = "technical"  # Pre-tender estimate
    TENDER = "tender"  # Tender pricing


class Region(Enum):
    """Regions for rate adjustment."""
    GAUTENG = "gauteng"
    WESTERN_CAPE = "western_cape"
    KWAZULU_NATAL = "kwazulu_natal"
    EASTERN_CAPE = "eastern_cape"
    FREE_STATE = "free_state"
    LIMPOPO = "limpopo"
    MPUMALANGA = "mpumalanga"
    NORTH_WEST = "north_west"
    NORTHERN_CAPE = "northern_cape"
    UK = "uk"


class RiskLevel(Enum):
    """Risk level for contingency calculation."""
    LOW = "low"
    MEDIUM = "medium"
    HIGH = "high"
    VERY_HIGH = "very_high"


@dataclass
class RateItem:
    """A rate from the rate database."""
    rate_id: str
    element_code: str  # NRM element code
    description: str
    rate: Decimal
    unit: UnitOfMeasure
    rate_source: str  # e.g., "BuildSmart Q3 2024", "Historical Project"
    rate_date: Optional[str] = None
    notes: Optional[str] = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "rate_id": self.rate_id,
            "element_code": self.element_code,
            "description": self.description,
            "rate": float(self.rate),
            "unit": self.unit.value,
            "rate_source": self.rate_source,
            "rate_date": self.rate_date,
            "notes": self.notes,
        }


@dataclass
class PricedItem:
    """A priced quantity item."""
    item_id: str
    quantity_ref: str  # Reference to source QuantityItem
    element_ref: str
    description: str
    quantity: Decimal
    unit: UnitOfMeasure
    rate: Decimal
    rate_source: str
    amount: Decimal
    regional_factor: Decimal = Decimal("1.0")
    adjusted_amount: Decimal = Decimal("0.0")
    confidence: float = 0.0
    notes: list[str] = field(default_factory=list)

    def __post_init__(self):
        """Calculate adjusted amount after init."""
        if self.adjusted_amount == Decimal("0.0"):
            self.adjusted_amount = self.amount * self.regional_factor

    def to_dict(self) -> dict[str, Any]:
        return {
            "item_id": self.item_id,
            "quantity_ref": self.quantity_ref,
            "element_ref": self.element_ref,
            "description": self.description,
            "quantity": float(self.quantity),
            "unit": self.unit.value,
            "rate": float(self.rate),
            "rate_source": self.rate_source,
            "amount": float(self.amount),
            "regional_factor": float(self.regional_factor),
            "adjusted_amount": float(self.adjusted_amount),
            "confidence": self.confidence,
            "notes": self.notes,
        }


@dataclass
class ElementCostGroup:
    """Group of priced items by element."""
    category: ElementCategory
    element_code: str
    element_name: str
    items: list[PricedItem] = field(default_factory=list)
    subtotal: Decimal = Decimal("0.0")

    def calculate_subtotal(self) -> None:
        """Calculate subtotal from items."""
        self.subtotal = sum(
            (item.adjusted_amount for item in self.items),
            Decimal("0.0")
        )

    def to_dict(self) -> dict[str, Any]:
        return {
            "category": self.category.value,
            "element_code": self.element_code,
            "element_name": self.element_name,
            "items": [i.to_dict() for i in self.items],
            "item_count": len(self.items),
            "subtotal": float(self.subtotal),
        }


@dataclass
class CostAllowance:
    """An allowance/addition to base cost."""
    allowance_id: str
    description: str
    percentage: Optional[Decimal] = None
    fixed_amount: Optional[Decimal] = None
    calculated_amount: Decimal = Decimal("0.0")
    basis: str = ""  # What it's calculated on
    notes: Optional[str] = None

    def calculate(self, base: Decimal) -> Decimal:
        """Calculate the allowance amount."""
        if self.fixed_amount is not None:
            self.calculated_amount = self.fixed_amount
        elif self.percentage is not None:
            self.calculated_amount = base * (self.percentage / Decimal("100"))
        return self.calculated_amount

    def to_dict(self) -> dict[str, Any]:
        return {
            "allowance_id": self.allowance_id,
            "description": self.description,
            "percentage": float(self.percentage) if self.percentage else None,
            "fixed_amount": float(self.fixed_amount) if self.fixed_amount else None,
            "calculated_amount": float(self.calculated_amount),
            "basis": self.basis,
            "notes": self.notes,
        }


@dataclass
class CostRisk:
    """A cost risk item."""
    risk_id: str
    category: str
    description: str
    likelihood: str  # "low", "medium", "high"
    impact: str  # "low", "medium", "high"
    cost_impact_low: Decimal = Decimal("0.0")
    cost_impact_high: Decimal = Decimal("0.0")
    mitigation: Optional[str] = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "risk_id": self.risk_id,
            "category": self.category,
            "description": self.description,
            "likelihood": self.likelihood,
            "impact": self.impact,
            "cost_impact_low": float(self.cost_impact_low),
            "cost_impact_high": float(self.cost_impact_high),
            "mitigation": self.mitigation,
        }


@dataclass
class CostSummary:
    """Summary of costs by category."""
    base_building_cost: Decimal = Decimal("0.0")
    external_works: Decimal = Decimal("0.0")
    preliminaries: Decimal = Decimal("0.0")
    contingencies: Decimal = Decimal("0.0")
    professional_fees: Decimal = Decimal("0.0")
    other_allowances: Decimal = Decimal("0.0")
    subtotal_excl_vat: Decimal = Decimal("0.0")
    vat_amount: Decimal = Decimal("0.0")
    total_incl_vat: Decimal = Decimal("0.0")
    cost_per_sqm: Optional[Decimal] = None
    gifa: Optional[Decimal] = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "base_building_cost": float(self.base_building_cost),
            "external_works": float(self.external_works),
            "preliminaries": float(self.preliminaries),
            "contingencies": float(self.contingencies),
            "professional_fees": float(self.professional_fees),
            "other_allowances": float(self.other_allowances),
            "subtotal_excl_vat": float(self.subtotal_excl_vat),
            "vat_amount": float(self.vat_amount),
            "total_incl_vat": float(self.total_incl_vat),
            "cost_per_sqm": float(self.cost_per_sqm) if self.cost_per_sqm else None,
            "gifa": float(self.gifa) if self.gifa else None,
        }


@dataclass
class CostResult:
    """Complete result of cost pricing."""
    project_id: str
    pricing_date: str
    project_stage: ProjectStage
    region: Region
    currency: str = "ZAR"
    element_groups: list[ElementCostGroup] = field(default_factory=list)
    allowances: list[CostAllowance] = field(default_factory=list)
    risks: list[CostRisk] = field(default_factory=list)
    summary: CostSummary = field(default_factory=CostSummary)
    processing_time_ms: float = 0.0
    items_priced: int = 0
    items_unpriced: int = 0
    pricing_assumptions: list[str] = field(default_factory=list)
    exclusions: list[str] = field(default_factory=list)
    errors: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "project_id": self.project_id,
            "pricing_date": self.pricing_date,
            "project_stage": self.project_stage.value,
            "region": self.region.value,
            "currency": self.currency,
            "element_groups": [g.to_dict() for g in self.element_groups],
            "allowances": [a.to_dict() for a in self.allowances],
            "risks": [r.to_dict() for r in self.risks],
            "summary": self.summary.to_dict(),
            "statistics": {
                "items_priced": self.items_priced,
                "items_unpriced": self.items_unpriced,
                "element_group_count": len(self.element_groups),
                "risk_count": len(self.risks),
            },
            "pricing_assumptions": self.pricing_assumptions,
            "exclusions": self.exclusions,
            "processing_time_ms": self.processing_time_ms,
            "errors": self.errors,
            "warnings": self.warnings,
        }

    def to_json(self, indent: int = 2) -> str:
        return json.dumps(self.to_dict(), indent=indent, default=str)


class CostAgent:
    """
    Cost Agent - Cost Planner applying rates and producing priced documents.

    Transforms quantities into money. Maintains rate databases, applies
    regional adjustment factors, incorporates market conditions, and
    calculates preliminaries, contingencies, and professional fees.

    Principles:
    - Price to the appropriate level of detail for project stage
    - Show rate sources and adjustment logic
    - Apply contingencies appropriate to project risk profile
    - Separate base cost from allowances and contingencies
    - Present ranges, not false precision
    """

    # Regional adjustment factors (baseline: Gauteng = 1.00)
    REGIONAL_FACTORS = {
        Region.GAUTENG: Decimal("1.00"),
        Region.WESTERN_CAPE: Decimal("1.05"),
        Region.KWAZULU_NATAL: Decimal("0.98"),
        Region.EASTERN_CAPE: Decimal("0.95"),
        Region.FREE_STATE: Decimal("0.92"),
        Region.LIMPOPO: Decimal("0.90"),
        Region.MPUMALANGA: Decimal("0.93"),
        Region.NORTH_WEST: Decimal("0.91"),
        Region.NORTHERN_CAPE: Decimal("0.88"),
        Region.UK: Decimal("1.00"),  # Separate rate database
    }

    # Contingency ranges by project type and risk
    CONTINGENCY_RANGES = {
        "new_build_standard": (Decimal("5.0"), Decimal("8.0")),
        "new_build_complex": (Decimal("8.0"), Decimal("12.0")),
        "refurbishment": (Decimal("15.0"), Decimal("20.0")),
        "heritage": (Decimal("20.0"), Decimal("30.0")),
        "tender_review": (Decimal("3.0"), Decimal("5.0")),
        "variation": (Decimal("5.0"), Decimal("10.0")),
    }

    # Preliminaries ranges by project size
    PRELIMINARIES_RANGES = {
        "small": (Decimal("8.0"), Decimal("12.0")),  # <R10m
        "medium": (Decimal("10.0"), Decimal("15.0")),  # R10m-R50m
        "large": (Decimal("12.0"), Decimal("18.0")),  # >R50m
    }

    # Professional fees by project stage
    PROFESSIONAL_FEES = {
        ProjectStage.FEASIBILITY: Decimal("1.0"),
        ProjectStage.CONCEPT: Decimal("2.0"),
        ProjectStage.DEVELOPED: Decimal("3.0"),
        ProjectStage.TECHNICAL: Decimal("4.0"),
        ProjectStage.TENDER: Decimal("5.0"),
    }

    # VAT rate
    VAT_RATE = Decimal("15.0")  # South Africa

    # Base rates by element code (R/unit) - simplified rate database
    # In production, this would come from a database or external source
    BASE_RATES = {
        # Substructure (1.x)
        "1.1": {"description": "Standard foundations", "rate": Decimal("850"), "unit": UnitOfMeasure.SQUARE_METRES},
        "1.1.1": {"description": "Strip foundations", "rate": Decimal("750"), "unit": UnitOfMeasure.LINEAR_METRES},
        "1.1.2": {"description": "Pad foundations", "rate": Decimal("2500"), "unit": UnitOfMeasure.CUBIC_METRES},
        "1.1.3": {"description": "Basement excavation", "rate": Decimal("180"), "unit": UnitOfMeasure.CUBIC_METRES},
        "1.1.4": {"description": "Basement retaining walls", "rate": Decimal("3500"), "unit": UnitOfMeasure.SQUARE_METRES},

        # Superstructure - Frame (2.1)
        "2.1": {"description": "Structural frame", "rate": Decimal("1800"), "unit": UnitOfMeasure.SQUARE_METRES},

        # Superstructure - Upper floors (2.2)
        "2.2": {"description": "Upper floor construction", "rate": Decimal("1200"), "unit": UnitOfMeasure.SQUARE_METRES},

        # Superstructure - Roof (2.3)
        "2.3": {"description": "Roof construction", "rate": Decimal("950"), "unit": UnitOfMeasure.SQUARE_METRES},

        # Superstructure - Stairs (2.4)
        "2.4": {"description": "Stairs and ramps", "rate": Decimal("15000"), "unit": UnitOfMeasure.NUMBER},

        # Superstructure - External walls (2.5)
        "2.5": {"description": "External walls", "rate": Decimal("1800"), "unit": UnitOfMeasure.SQUARE_METRES},

        # Windows and external doors (2.6)
        "2.6": {"description": "Windows and external doors", "rate": Decimal("3500"), "unit": UnitOfMeasure.SQUARE_METRES},

        # Internal walls (2.7)
        "2.7": {"description": "Internal walls and partitions", "rate": Decimal("650"), "unit": UnitOfMeasure.SQUARE_METRES},

        # Internal doors (2.8)
        "2.8": {"description": "Internal doors", "rate": Decimal("4500"), "unit": UnitOfMeasure.NUMBER},

        # Internal finishes - Walls (3.1)
        "3.1": {"description": "Wall finishes", "rate": Decimal("280"), "unit": UnitOfMeasure.SQUARE_METRES},

        # Internal finishes - Floors (3.2)
        "3.2": {"description": "Floor finishes", "rate": Decimal("450"), "unit": UnitOfMeasure.SQUARE_METRES},

        # Internal finishes - Ceilings (3.3)
        "3.3": {"description": "Ceiling finishes", "rate": Decimal("320"), "unit": UnitOfMeasure.SQUARE_METRES},

        # Services (5.x)
        "5.1": {"description": "Sanitary installations", "rate": Decimal("18000"), "unit": UnitOfMeasure.NUMBER},
        "5.4": {"description": "Water installations", "rate": Decimal("85"), "unit": UnitOfMeasure.SQUARE_METRES},
        "5.6": {"description": "HVAC", "rate": Decimal("350"), "unit": UnitOfMeasure.SQUARE_METRES},
        "5.8": {"description": "Electrical installations", "rate": Decimal("280"), "unit": UnitOfMeasure.SQUARE_METRES},

        # External works (8.x)
        "8.1": {"description": "Site preparation", "rate": Decimal("120"), "unit": UnitOfMeasure.SQUARE_METRES},
        "8.2": {"description": "Roads, paths, pavings", "rate": Decimal("450"), "unit": UnitOfMeasure.SQUARE_METRES},
        "8.3": {"description": "Soft landscaping", "rate": Decimal("180"), "unit": UnitOfMeasure.SQUARE_METRES},
        "8.4": {"description": "Fencing, railings, walls", "rate": Decimal("850"), "unit": UnitOfMeasure.LINEAR_METRES},
        "8.5": {"description": "External drainage", "rate": Decimal("650"), "unit": UnitOfMeasure.LINEAR_METRES},

        # Default/fallback
        "9.0": {"description": "Miscellaneous", "rate": Decimal("500"), "unit": UnitOfMeasure.ITEM},
    }

    def __init__(self, config: Optional[dict[str, Any]] = None):
        """
        Initialize the Cost Agent.

        Args:
            config: Configuration dict with optional keys:
                - output_dir: Directory for output files
                - region: Project region for rate adjustment
                - project_stage: Pricing detail level
                - project_type: For contingency calculation
                - currency: Currency code (default ZAR)
                - vat_rate: VAT percentage (default 15.0)
                - custom_rates: Path to custom rate database
        """
        self.config = config or {}
        self.output_dir = Path(self.config.get("output_dir", "./cost_output"))

        # Region for adjustment factors
        region_str = self.config.get("region", "gauteng")
        try:
            self.region = Region(region_str.lower())
        except ValueError:
            self.region = Region.GAUTENG
            logger.warning(f"Unknown region '{region_str}', defaulting to Gauteng")

        # Project stage
        stage_str = self.config.get("project_stage", "concept")
        try:
            self.project_stage = ProjectStage(stage_str.lower())
        except ValueError:
            self.project_stage = ProjectStage.CONCEPT

        # Project type for contingency
        self.project_type = self.config.get("project_type", "new_build_standard")

        # Currency
        self.currency = self.config.get("currency", "ZAR")

        # VAT
        vat = self.config.get("vat_rate")
        self.vat_rate = Decimal(str(vat)) if vat else self.VAT_RATE

        # Load custom rates if provided
        self.rates = dict(self.BASE_RATES)
        if self.config.get("custom_rates"):
            self._load_custom_rates(self.config["custom_rates"])

        self.logger = logging.getLogger(self.__class__.__name__)

        # Item counter
        self._item_counter = 0

    def _generate_item_id(self) -> str:
        """Generate unique item ID."""
        self._item_counter += 1
        return f"COST-{self._item_counter:04d}"

    def _generate_risk_id(self) -> str:
        """Generate unique risk ID."""
        return f"RISK-{uuid.uuid4().hex[:6].upper()}"

    def _generate_allowance_id(self) -> str:
        """Generate unique allowance ID."""
        return f"ALW-{uuid.uuid4().hex[:6].upper()}"

    def _load_custom_rates(self, rates_path: str) -> None:
        """Load custom rates from file."""
        try:
            with open(rates_path) as f:
                custom = json.load(f)
                for code, data in custom.items():
                    self.rates[code] = {
                        "description": data.get("description", ""),
                        "rate": Decimal(str(data.get("rate", 0))),
                        "unit": UnitOfMeasure(data.get("unit", "item")),
                    }
            self.logger.info(f"Loaded {len(custom)} custom rates from {rates_path}")
        except Exception as e:
            self.logger.warning(f"Failed to load custom rates: {e}")

    async def price(
        self,
        project_id: str,
        measure_result: MeasureResult,
        gifa: Optional[float] = None,
    ) -> CostResult:
        """
        Apply rates to quantities and produce priced documents.

        Args:
            project_id: Project identifier
            measure_result: Results from Measure Agent
            gifa: Gross Internal Floor Area for cost/m² calculation

        Returns:
            CostResult with priced BOQ, summary, and risks
        """
        start_time = time.time()

        self.output_dir.mkdir(parents=True, exist_ok=True)
        self._item_counter = 0

        errors: list[dict[str, Any]] = []
        warnings: list[str] = []
        pricing_assumptions: list[str] = []

        self.logger.info(f"[{project_id}] Starting cost pricing")
        self.logger.info(f"[{project_id}] Region: {self.region.value}, Stage: {self.project_stage.value}")

        # Get regional adjustment factor
        regional_factor = self.REGIONAL_FACTORS.get(self.region, Decimal("1.0"))
        pricing_assumptions.append(f"Regional adjustment factor: {regional_factor} ({self.region.value})")

        # Price each element group
        priced_groups: list[ElementCostGroup] = []
        items_priced = 0
        items_unpriced = 0

        for group in measure_result.element_groups:
            priced_group = ElementCostGroup(
                category=group.category,
                element_code=group.element_code,
                element_name=group.element_name,
            )

            for qty_item in group.items:
                priced_item = self._price_item(qty_item, regional_factor)

                if priced_item:
                    priced_group.items.append(priced_item)
                    items_priced += 1
                else:
                    items_unpriced += 1
                    warnings.append(f"No rate found for: {qty_item.description} ({qty_item.element_ref})")

            priced_group.calculate_subtotal()
            if priced_group.items:
                priced_groups.append(priced_group)

        self.logger.info(f"[{project_id}] Priced {items_priced} items, {items_unpriced} unpriced")

        # Calculate base costs
        base_building_cost = sum(
            (g.subtotal for g in priced_groups if g.category != ElementCategory.EXTERNAL_WORKS),
            Decimal("0.0")
        )
        external_works = sum(
            (g.subtotal for g in priced_groups if g.category == ElementCategory.EXTERNAL_WORKS),
            Decimal("0.0")
        )

        # Calculate allowances
        allowances: list[CostAllowance] = []

        # Preliminaries
        prelim_pct = self._calculate_preliminaries_percentage(base_building_cost)
        preliminaries = CostAllowance(
            allowance_id=self._generate_allowance_id(),
            description="Preliminaries",
            percentage=prelim_pct,
            basis="Base building cost",
            notes=self._get_preliminaries_notes(),
        )
        preliminaries.calculate(base_building_cost)
        allowances.append(preliminaries)
        pricing_assumptions.append(f"Preliminaries: {prelim_pct}% based on project size")

        # Contingencies
        contingency_pct = self._calculate_contingency_percentage()
        contingencies = CostAllowance(
            allowance_id=self._generate_allowance_id(),
            description="Contingencies",
            percentage=contingency_pct,
            basis="Base building cost + external works",
            notes=f"Based on {self.project_type} project type",
        )
        contingencies.calculate(base_building_cost + external_works)
        allowances.append(contingencies)
        pricing_assumptions.append(f"Contingency: {contingency_pct}% for {self.project_type}")

        # Professional fees (if applicable at this stage)
        fees_pct = self.PROFESSIONAL_FEES.get(self.project_stage, Decimal("0.0"))
        if fees_pct > 0:
            fees = CostAllowance(
                allowance_id=self._generate_allowance_id(),
                description="Professional fees allowance",
                percentage=fees_pct,
                basis="Construction cost",
                notes=f"Allowance for {self.project_stage.value} stage",
            )
            subtotal_for_fees = base_building_cost + external_works + preliminaries.calculated_amount
            fees.calculate(subtotal_for_fees)
            allowances.append(fees)
            pricing_assumptions.append(f"Professional fees: {fees_pct}% allowance")

        # Build cost summary
        summary = self._build_cost_summary(
            base_building_cost=base_building_cost,
            external_works=external_works,
            allowances=allowances,
            gifa=Decimal(str(gifa)) if gifa else None,
        )

        # Identify risks
        risks = self._identify_risks(measure_result, summary)

        # Determine exclusions
        exclusions = self._determine_exclusions(measure_result)

        processing_time = (time.time() - start_time) * 1000

        result = CostResult(
            project_id=project_id,
            pricing_date=datetime.now().isoformat(),
            project_stage=self.project_stage,
            region=self.region,
            currency=self.currency,
            element_groups=priced_groups,
            allowances=allowances,
            risks=risks,
            summary=summary,
            processing_time_ms=processing_time,
            items_priced=items_priced,
            items_unpriced=items_unpriced,
            pricing_assumptions=pricing_assumptions,
            exclusions=exclusions,
            errors=errors,
            warnings=warnings,
        )

        # Save outputs
        await self._save_outputs(project_id, result)

        self.logger.info(f"[{project_id}] Cost pricing complete: {self.currency} {summary.total_incl_vat:,.2f}")

        return result

    def _price_item(
        self,
        qty_item: QuantityItem,
        regional_factor: Decimal,
    ) -> Optional[PricedItem]:
        """Price a single quantity item."""
        # Find matching rate
        rate_data = self._find_rate(qty_item.element_ref, qty_item.unit)

        if not rate_data:
            return None

        rate = rate_data["rate"]
        quantity = Decimal(str(qty_item.quantity))
        amount = quantity * rate
        adjusted_amount = amount * regional_factor

        return PricedItem(
            item_id=self._generate_item_id(),
            quantity_ref=qty_item.item_id,
            element_ref=qty_item.element_ref,
            description=qty_item.description,
            quantity=quantity,
            unit=qty_item.unit,
            rate=rate,
            rate_source=f"Base rates Q4 2024 ({rate_data['description']})",
            amount=amount,
            regional_factor=regional_factor,
            adjusted_amount=adjusted_amount,
            confidence=qty_item.confidence,
            notes=qty_item.notes.copy() if qty_item.notes else [],
        )

    def _find_rate(
        self,
        element_ref: str,
        unit: UnitOfMeasure,
    ) -> Optional[dict[str, Any]]:
        """Find the best matching rate for an element."""
        # Try exact match first
        if element_ref in self.rates:
            rate_data = self.rates[element_ref]
            if rate_data["unit"] == unit or rate_data["unit"].value == unit.value:
                return rate_data

        # Try parent element code (e.g., "2.5.1" -> "2.5" -> "2")
        parts = element_ref.split(".")
        for i in range(len(parts) - 1, 0, -1):
            parent_ref = ".".join(parts[:i])
            if parent_ref in self.rates:
                return self.rates[parent_ref]

        # Fallback to default rate
        if "9.0" in self.rates:
            return self.rates["9.0"]

        return None

    def _calculate_preliminaries_percentage(self, base_cost: Decimal) -> Decimal:
        """Calculate preliminaries percentage based on project size."""
        # Determine project size category
        if base_cost < Decimal("10000000"):  # <R10m
            low, high = self.PRELIMINARIES_RANGES["small"]
        elif base_cost < Decimal("50000000"):  # R10m-R50m
            low, high = self.PRELIMINARIES_RANGES["medium"]
        else:  # >R50m
            low, high = self.PRELIMINARIES_RANGES["large"]

        # Use midpoint for estimate
        return (low + high) / 2

    def _get_preliminaries_notes(self) -> str:
        """Get notes about preliminaries calculation."""
        return (
            "Includes site establishment, management, insurance, temporary works, "
            "plant, scaffolding, and site overheads"
        )

    def _calculate_contingency_percentage(self) -> Decimal:
        """Calculate contingency percentage based on project type."""
        range_key = self.project_type.lower().replace(" ", "_")

        if range_key in self.CONTINGENCY_RANGES:
            low, high = self.CONTINGENCY_RANGES[range_key]
        else:
            # Default to standard new build
            low, high = self.CONTINGENCY_RANGES["new_build_standard"]

        # Use midpoint for estimate
        return (low + high) / 2

    def _build_cost_summary(
        self,
        base_building_cost: Decimal,
        external_works: Decimal,
        allowances: list[CostAllowance],
        gifa: Optional[Decimal] = None,
    ) -> CostSummary:
        """Build the cost summary."""
        # Extract allowance amounts
        preliminaries = Decimal("0.0")
        contingencies = Decimal("0.0")
        professional_fees = Decimal("0.0")
        other_allowances = Decimal("0.0")

        for a in allowances:
            if "preliminaries" in a.description.lower():
                preliminaries = a.calculated_amount
            elif "contingenc" in a.description.lower():
                contingencies = a.calculated_amount
            elif "professional" in a.description.lower() or "fees" in a.description.lower():
                professional_fees = a.calculated_amount
            else:
                other_allowances += a.calculated_amount

        # Calculate totals
        subtotal = (
            base_building_cost +
            external_works +
            preliminaries +
            contingencies +
            professional_fees +
            other_allowances
        )

        vat_amount = subtotal * (self.vat_rate / Decimal("100"))
        total = subtotal + vat_amount

        # Cost per m²
        cost_per_sqm = None
        if gifa and gifa > 0:
            cost_per_sqm = total / gifa

        return CostSummary(
            base_building_cost=base_building_cost.quantize(Decimal("0.01"), ROUND_HALF_UP),
            external_works=external_works.quantize(Decimal("0.01"), ROUND_HALF_UP),
            preliminaries=preliminaries.quantize(Decimal("0.01"), ROUND_HALF_UP),
            contingencies=contingencies.quantize(Decimal("0.01"), ROUND_HALF_UP),
            professional_fees=professional_fees.quantize(Decimal("0.01"), ROUND_HALF_UP),
            other_allowances=other_allowances.quantize(Decimal("0.01"), ROUND_HALF_UP),
            subtotal_excl_vat=subtotal.quantize(Decimal("0.01"), ROUND_HALF_UP),
            vat_amount=vat_amount.quantize(Decimal("0.01"), ROUND_HALF_UP),
            total_incl_vat=total.quantize(Decimal("0.01"), ROUND_HALF_UP),
            cost_per_sqm=cost_per_sqm.quantize(Decimal("0.01"), ROUND_HALF_UP) if cost_per_sqm else None,
            gifa=gifa,
        )

    def _identify_risks(
        self,
        measure_result: MeasureResult,
        summary: CostSummary,
    ) -> list[CostRisk]:
        """Identify cost risks based on measurement and pricing."""
        risks: list[CostRisk] = []

        # Risk: Low confidence measurements
        low_conf_items = []
        for group in measure_result.element_groups:
            for item in group.items:
                if item.confidence < 0.5:
                    low_conf_items.append(item)

        if low_conf_items:
            impact_pct = len(low_conf_items) / max(measure_result.total_items, 1)
            risks.append(CostRisk(
                risk_id=self._generate_risk_id(),
                category="measurement_uncertainty",
                description=f"{len(low_conf_items)} items with low measurement confidence",
                likelihood="medium",
                impact="medium" if impact_pct < 0.2 else "high",
                cost_impact_low=summary.base_building_cost * Decimal("0.02"),
                cost_impact_high=summary.base_building_cost * Decimal("0.08"),
                mitigation="Obtain clearer drawings or site verification for low-confidence items",
            ))

        # Risk: Clarifications outstanding
        if measure_result.clarifications:
            high_priority = [c for c in measure_result.clarifications if c.priority == "high"]
            risks.append(CostRisk(
                risk_id=self._generate_risk_id(),
                category="design_uncertainty",
                description=f"{len(measure_result.clarifications)} clarifications outstanding ({len(high_priority)} high priority)",
                likelihood="high" if high_priority else "medium",
                impact="high" if high_priority else "medium",
                cost_impact_low=summary.base_building_cost * Decimal("0.03"),
                cost_impact_high=summary.base_building_cost * Decimal("0.10"),
                mitigation="Resolve design clarifications before proceeding to tender",
            ))

        # Risk: Market volatility
        risks.append(CostRisk(
            risk_id=self._generate_risk_id(),
            category="market",
            description="Material price volatility",
            likelihood="medium",
            impact="medium",
            cost_impact_low=summary.base_building_cost * Decimal("0.02"),
            cost_impact_high=summary.base_building_cost * Decimal("0.05"),
            mitigation="Consider price escalation clauses and early procurement of key materials",
        ))

        # Risk: Scope changes
        risks.append(CostRisk(
            risk_id=self._generate_risk_id(),
            category="scope",
            description="Potential scope changes during construction",
            likelihood="medium",
            impact="medium",
            cost_impact_low=summary.base_building_cost * Decimal("0.03"),
            cost_impact_high=summary.base_building_cost * Decimal("0.10"),
            mitigation="Thorough design review and client sign-off before tender",
        ))

        return risks

    def _determine_exclusions(self, measure_result: MeasureResult) -> list[str]:
        """Determine pricing exclusions."""
        exclusions = list(measure_result.exclusions)

        # Add standard pricing exclusions
        standard_exclusions = [
            "Land acquisition costs",
            "Finance charges and interest",
            "Loose furniture and equipment",
            "Client direct works",
            "Abnormal ground conditions (provisional)",
        ]

        for excl in standard_exclusions:
            if excl not in exclusions:
                exclusions.append(excl)

        return exclusions

    async def _save_outputs(self, project_id: str, result: CostResult) -> None:
        """Save output files."""
        project_dir = self.output_dir
        project_dir.mkdir(parents=True, exist_ok=True)

        # Save priced_boq.json (XLSX generation would require openpyxl)
        boq_path = project_dir / "priced_boq.json"
        with open(boq_path, "w") as f:
            f.write(result.to_json(indent=2))
        self.logger.info(f"Saved: {boq_path}")

        # Generate XLSX if openpyxl available
        try:
            await self._save_xlsx(project_dir / "priced_boq.xlsx", result)
        except ImportError:
            self.logger.warning("openpyxl not installed - skipping XLSX output")

        # Save cost_summary.md
        summary_path = project_dir / "cost_summary.md"
        with open(summary_path, "w") as f:
            f.write(self._generate_cost_summary_md(project_id, result))
        self.logger.info(f"Saved: {summary_path}")

        # Save pricing_assumptions.md
        assumptions_path = project_dir / "pricing_assumptions.md"
        with open(assumptions_path, "w") as f:
            f.write(self._generate_assumptions_md(project_id, result))
        self.logger.info(f"Saved: {assumptions_path}")

        # Save cost_risk_register.md
        risks_path = project_dir / "cost_risk_register.md"
        with open(risks_path, "w") as f:
            f.write(self._generate_risks_md(project_id, result))
        self.logger.info(f"Saved: {risks_path}")

    async def _save_xlsx(self, path: Path, result: CostResult) -> None:
        """Save priced BOQ as Excel file."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Header style
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        currency_format = f'#,##0.00 "{result.currency}"'

        # Summary content
        summary_data = [
            ["COST SUMMARY", ""],
            ["", ""],
            ["Project ID", result.project_id],
            ["Pricing Date", result.pricing_date[:10]],
            ["Region", result.region.value.replace("_", " ").title()],
            ["Project Stage", result.project_stage.value.title()],
            ["", ""],
            ["COST BREAKDOWN", "Amount"],
            ["Base Building Cost", float(result.summary.base_building_cost)],
            ["External Works", float(result.summary.external_works)],
            ["Preliminaries", float(result.summary.preliminaries)],
            ["Contingencies", float(result.summary.contingencies)],
            ["Professional Fees", float(result.summary.professional_fees)],
            ["Other Allowances", float(result.summary.other_allowances)],
            ["", ""],
            ["SUBTOTAL (excl VAT)", float(result.summary.subtotal_excl_vat)],
            ["VAT @ 15%", float(result.summary.vat_amount)],
            ["TOTAL (incl VAT)", float(result.summary.total_incl_vat)],
        ]

        if result.summary.cost_per_sqm:
            summary_data.extend([
                ["", ""],
                ["GIFA", f"{float(result.summary.gifa):.2f} m²"],
                ["Cost per m²", float(result.summary.cost_per_sqm)],
            ])

        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1 or row_idx == 8:
                    cell.font = header_font
                    cell.fill = header_fill
                if col_idx == 2 and isinstance(value, (int, float)) and row_idx > 7:
                    cell.number_format = currency_format

        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20

        # BOQ sheet
        ws_boq = wb.create_sheet("Priced BOQ")

        boq_headers = ["Item", "Element", "Description", "Qty", "Unit", "Rate", "Amount", "Notes"]
        for col_idx, header in enumerate(boq_headers, 1):
            cell = ws_boq.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row_idx = 2
        for group in result.element_groups:
            # Group header
            ws_boq.cell(row=row_idx, column=1, value=group.element_code)
            cell = ws_boq.cell(row=row_idx, column=3, value=group.element_name.upper())
            cell.font = header_font
            row_idx += 1

            for item in group.items:
                ws_boq.cell(row=row_idx, column=1, value=item.item_id)
                ws_boq.cell(row=row_idx, column=2, value=item.element_ref)
                ws_boq.cell(row=row_idx, column=3, value=item.description)
                ws_boq.cell(row=row_idx, column=4, value=float(item.quantity))
                ws_boq.cell(row=row_idx, column=5, value=item.unit.value)
                ws_boq.cell(row=row_idx, column=6, value=float(item.rate)).number_format = currency_format
                ws_boq.cell(row=row_idx, column=7, value=float(item.adjusted_amount)).number_format = currency_format
                ws_boq.cell(row=row_idx, column=8, value="; ".join(item.notes) if item.notes else "")
                row_idx += 1

            # Group subtotal
            ws_boq.cell(row=row_idx, column=3, value="Subtotal")
            cell = ws_boq.cell(row=row_idx, column=7, value=float(group.subtotal))
            cell.number_format = currency_format
            cell.font = header_font
            row_idx += 2

        # Set column widths
        col_widths = [10, 10, 40, 10, 8, 15, 18, 30]
        for col_idx, width in enumerate(col_widths, 1):
            ws_boq.column_dimensions[get_column_letter(col_idx)].width = width

        wb.save(path)
        self.logger.info(f"Saved: {path}")

    def _generate_cost_summary_md(self, project_id: str, result: CostResult) -> str:
        """Generate cost summary markdown."""
        s = result.summary
        lines = [
            "# Cost Summary",
            "",
            f"**Project ID:** {project_id}",
            f"**Pricing Date:** {result.pricing_date[:10]}",
            f"**Region:** {result.region.value.replace('_', ' ').title()}",
            f"**Project Stage:** {result.project_stage.value.title()}",
            f"**Currency:** {result.currency}",
            "",
            "---",
            "",
            "## Cost Breakdown",
            "",
            "| Category | Amount |",
            "|----------|-------:|",
            f"| Base Building Cost | {s.base_building_cost:,.2f} |",
            f"| External Works | {s.external_works:,.2f} |",
            f"| Preliminaries | {s.preliminaries:,.2f} |",
            f"| Contingencies | {s.contingencies:,.2f} |",
            f"| Professional Fees | {s.professional_fees:,.2f} |",
            f"| Other Allowances | {s.other_allowances:,.2f} |",
            "| | |",
            f"| **Subtotal (excl VAT)** | **{s.subtotal_excl_vat:,.2f}** |",
            f"| VAT @ {self.vat_rate}% | {s.vat_amount:,.2f} |",
            f"| **TOTAL (incl VAT)** | **{s.total_incl_vat:,.2f}** |",
            "",
        ]

        if s.cost_per_sqm and s.gifa:
            lines.extend([
                "---",
                "",
                "## Cost per Square Metre",
                "",
                f"- **GIFA:** {s.gifa:,.2f} m²",
                f"- **Cost/m²:** {result.currency} {s.cost_per_sqm:,.2f}",
                "",
            ])

        lines.extend([
            "---",
            "",
            "## Element Breakdown",
            "",
        ])

        for group in result.element_groups:
            lines.append(f"### {group.element_code} - {group.element_name}")
            lines.append("")
            lines.append(f"**Subtotal:** {result.currency} {group.subtotal:,.2f}")
            lines.append("")

            if group.items:
                lines.append("| Description | Qty | Unit | Amount |")
                lines.append("|-------------|----:|------|-------:|")
                for item in group.items[:10]:
                    lines.append(
                        f"| {item.description[:35]} | {item.quantity:.2f} | {item.unit.value} | {item.adjusted_amount:,.2f} |"
                    )
                if len(group.items) > 10:
                    lines.append(f"| *...and {len(group.items) - 10} more items* | | | |")
            lines.append("")

        # Statistics
        lines.extend([
            "---",
            "",
            "## Statistics",
            "",
            f"- Items priced: {result.items_priced}",
            f"- Items unpriced: {result.items_unpriced}",
            f"- Processing time: {result.processing_time_ms:.0f}ms",
            "",
        ])

        if result.warnings:
            lines.extend([
                "---",
                "",
                "## Warnings",
                "",
            ])
            for warning in result.warnings[:10]:
                lines.append(f"- {warning}")
            if len(result.warnings) > 10:
                lines.append(f"- *...and {len(result.warnings) - 10} more*")

        return "\n".join(lines)

    def _generate_assumptions_md(self, project_id: str, result: CostResult) -> str:
        """Generate pricing assumptions markdown."""
        lines = [
            "# Pricing Assumptions",
            "",
            f"**Project ID:** {project_id}",
            f"**Pricing Date:** {result.pricing_date[:10]}",
            "",
            "---",
            "",
            "## Rate Sources",
            "",
            "Rates applied from the following sources:",
            "",
            "1. **Base Rates Database** - Internal rate library (Q4 2024 baseline)",
            f"2. **Regional Factor** - {self.REGIONAL_FACTORS.get(result.region, Decimal('1.0'))} for {result.region.value.replace('_', ' ').title()}",
            "",
            "---",
            "",
            "## Pricing Assumptions",
            "",
        ]

        for assumption in result.pricing_assumptions:
            lines.append(f"- {assumption}")

        lines.extend([
            "",
            "---",
            "",
            "## Allowances",
            "",
            "| Allowance | Percentage | Amount | Basis |",
            "|-----------|----------:|-------:|-------|",
        ])

        for a in result.allowances:
            pct = f"{a.percentage:.1f}%" if a.percentage else "Fixed"
            lines.append(
                f"| {a.description} | {pct} | {a.calculated_amount:,.2f} | {a.basis} |"
            )

        lines.extend([
            "",
            "---",
            "",
            "## Exclusions",
            "",
            "The following items are **excluded** from this estimate:",
            "",
        ])

        for excl in result.exclusions:
            lines.append(f"- {excl}")

        lines.extend([
            "",
            "---",
            "",
            "## Notes",
            "",
            "1. All rates are exclusive of VAT unless stated otherwise",
            "2. Rates are based on competitive tender conditions",
            "3. Prices assume normal working hours and conditions",
            "4. Specialist works priced on provisional/PC sum basis where noted",
            "5. This estimate is valid for 90 days from the pricing date",
            "",
        ])

        return "\n".join(lines)

    def _generate_risks_md(self, project_id: str, result: CostResult) -> str:
        """Generate cost risk register markdown."""
        lines = [
            "# Cost Risk Register",
            "",
            f"**Project ID:** {project_id}",
            f"**Assessment Date:** {result.pricing_date[:10]}",
            "",
            "---",
            "",
            "## Risk Summary",
            "",
            f"Total risks identified: {len(result.risks)}",
            "",
        ]

        # Count by impact
        high_impact = [r for r in result.risks if r.impact == "high"]
        med_impact = [r for r in result.risks if r.impact == "medium"]
        low_impact = [r for r in result.risks if r.impact == "low"]

        lines.extend([
            f"- High impact: {len(high_impact)}",
            f"- Medium impact: {len(med_impact)}",
            f"- Low impact: {len(low_impact)}",
            "",
            "---",
            "",
            "## Risk Register",
            "",
        ])

        for risk in result.risks:
            lines.extend([
                f"### {risk.risk_id}: {risk.description}",
                "",
                f"**Category:** {risk.category.replace('_', ' ').title()}",
                "",
                f"| Likelihood | Impact | Cost Range |",
                f"|------------|--------|------------|",
                f"| {risk.likelihood.title()} | {risk.impact.title()} | {result.currency} {risk.cost_impact_low:,.0f} - {risk.cost_impact_high:,.0f} |",
                "",
            ])

            if risk.mitigation:
                lines.extend([
                    f"**Mitigation:** {risk.mitigation}",
                    "",
                ])

            lines.append("---")
            lines.append("")

        # Total risk exposure
        total_low = sum(r.cost_impact_low for r in result.risks)
        total_high = sum(r.cost_impact_high for r in result.risks)

        lines.extend([
            "## Total Risk Exposure",
            "",
            f"| Scenario | Amount |",
            f"|----------|-------:|",
            f"| Best case (all risks low) | {result.currency} {total_low:,.0f} |",
            f"| Worst case (all risks high) | {result.currency} {total_high:,.0f} |",
            "",
        ])

        return "\n".join(lines)


async def run_cost(
    project_id: str,
    measure_result: MeasureResult,
    gifa: Optional[float] = None,
    output_dir: str = "./cost_output",
    **kwargs,
) -> CostResult:
    """
    Convenience function to run cost pricing.

    Args:
        project_id: Project identifier
        measure_result: Results from Measure Agent
        gifa: Gross Internal Floor Area for cost/m² calculation
        output_dir: Output directory for results
        **kwargs: Additional config options

    Returns:
        CostResult
    """
    config = {
        "output_dir": output_dir,
        **kwargs,
    }

    agent = CostAgent(config)
    return await agent.price(
        project_id=project_id,
        measure_result=measure_result,
        gifa=gifa,
    )
